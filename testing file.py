import paramiko
import time
import concurrent.futures
import json
import threading
from pathlib import Path

with open('sshsnmp_config.json', 'r') as config_file:
    data = json.load(config_file)

snmp_community = data["snmp_community"]
#ssh_lock = threading.Lock() #lock para controlar o acesso à conexão SSH


def ssh_connect(ssh_host=data["ssh_host"], ssh_username=data["ssh_username"], ssh_password=data["ssh_password"]):
    trys = 3
    while trys > 0:
        try:
            ssh_client = paramiko.SSHClient()
            ssh_client.set_missing_host_key_policy(
                paramiko.AutoAddPolicy())  # Adiciona chave de host, caso seja a primeira vez conectando
            ssh_client.connect(ssh_host, username=ssh_username, password=ssh_password)
            print("conectado ao SSH")
            return ssh_client

        except paramiko.ssh_exception.SSHException as e:
            trys -= 1
            time.sleep(3)
            if trys == 0:
                return f'Erro ao abrir canal SSH', ''


# tools
def extract_xlxs():
    from openpyxl import load_workbook

    workbook = load_workbook('C:/Users/INSS/Documents/hosts com falha SNMP.xlsx')
    sheet = workbook['Sheet']
    xlsx_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[2]:
            ip = row[1]
            host_name = row[0]
            xlsx_list.append({'host': host_name, 'ip': ip})  # Extend para adicionar elementos individualmente
    return xlsx_list


def save_excel(host_list, qnt=0):
    from openpyxl import Workbook

    path = Path('C:/Users/INSS/Documents/log SNMPwalk MAC-PORTA.xlsx')
    wb = Workbook()
    hosts_xlsx = wb.active
    count = 1

    hosts_xlsx[f'A{count}'] = 'hostname'
    hosts_xlsx[f'B{count}'] = 'IP'
    hosts_xlsx[f'C{count}'] = 'MAC'
    hosts_xlsx[f'D{count}'] = 'PORTA'

    count += 1

    # for host in host_list:
    #     # city_str = ', '.join(host['city'])
    #     hosts_xlsx[f'A{count}'] = host['name']
    #     hosts_xlsx[f'B{count}'] = host['ip']
    #     hosts_xlsx[f'C{count}'] = host['MAC']
    #     hosts_xlsx[f'D{count}'] = host['PORTA']
    #     count += 1
    for host, ip, mac_port_data in host_list:
        try:
            mac_list = mac_port_data['mac']
            port_list = mac_port_data['port']
            # if 'No Such Instance currently exists at this OID' in mac_list or port_list:
            #     mac = 'No Such Instance currently exists at this OID'
            #     port = 'No Such Instance currently exists at this OID'
            #     hosts_xlsx[f'A{count}'] = host
            #     hosts_xlsx[f'B{count}'] = ip
            #     hosts_xlsx[f'C{count}'] = mac
            #     hosts_xlsx[f'D{count}'] = port
            #     count += 1
            #     continue  # Vá para o próximo host

            valid_entries = [(mac.strip(), port.strip()) for mac, port in zip(mac_list, port_list) if
                             mac.strip() and port.strip()]

            if not valid_entries:  # Se não houver entradas válidas
                mac = 'Não foi possível acesso SNMP ao IP.'
                port = 'Não foi possível acesso SNMP ao IP.'
                hosts_xlsx[f'A{count}'] = host
                hosts_xlsx[f'B{count}'] = ip
                hosts_xlsx[f'C{count}'] = mac
                hosts_xlsx[f'D{count}'] = port
                count += 1
                continue  # Vá para o próximo host

            for mac_info, port_info in valid_entries:
                mac = mac_info.split("Hex-STRING: ")[1]
                # print(mac)
                port = port_info.split("INTEGER: ")[1]
                # print(port)
                hosts_xlsx[f'A{count}'] = host
                hosts_xlsx[f'B{count}'] = ip
                hosts_xlsx[f'C{count}'] = mac
                hosts_xlsx[f'D{count}'] = port
                count += 1
        except IndexError:
            mac = 'No Such Instance currently exists at this OID'
            port = 'No Such Instance currently exists at this OID'
            hosts_xlsx[f'A{count}'] = host
            hosts_xlsx[f'B{count}'] = ip
            hosts_xlsx[f'C{count}'] = mac
            hosts_xlsx[f'D{count}'] = port
            count += 1
            continue

    # hosts_xlsx['F1'] = 'Quantidade total de switches em falha SNMP'
    # hosts_xlsx['F2'] = qnt
    hosts_xlsx['E1'] = f'Gerada as {time.strftime("%H:%M:%S")}'

    wb.save(path)
    print(f'salvo em {path}')


def snmpwalk(snmp_host, oid, oid2, ssh_client,thread_lock,trys = 3):
    with thread_lock:
        while trys > 0:
            try:
                snmp_command_mac = f'snmpwalk -v 2c -c {snmp_community} {snmp_host} {oid}'
                snmp_command_port = f'snmpwalk -v 2c -c {snmp_community} {snmp_host} {oid2}'

                stdin, stdout, stderr = ssh_client.exec_command(snmp_command_mac)
                stdin2, stdout2, stderr2 = ssh_client.exec_command(snmp_command_port)
                print(f'executao: {snmp_command_mac}')
                output = stdout.read().decode()
                error = stderr.read().decode()
                # time.sleep(60)

                print(f'executao: {snmp_command_port}')
                output2 = stdout2.read().decode()
                error2 = stderr2.read().decode()

                if error and 'Could not chdir to home directory' not in error:
                    return error, error2
                if output and output2:
                    trys = 3
                    return output, output2
                if output == '':
                    trys = 3
                    return output, output2

            except concurrent.futures.TimeoutError:
                trys -= 1
                if trys == 0:
                    return f'Timeout ao obter resultado para {snmp_host}', ''
                else:
                    snmpwalk(snmp_host, oid, oid2, ssh_client, thread_lock, trys)
            except BaseException as e:
                trys -= 1
                if trys == 0:
                    return f'Erro ao conectar com o host {snmp_host} {e}', ''
                else:
                    snmpwalk(snmp_host, oid, oid2, ssh_client, thread_lock, trys)
    pass


def thread_execute(n_threads = 30, connection_lock = 30):
    thread_lock = threading.Semaphore(connection_lock)  # semáforo para controlar o número máximo de threads ativas
    ##teste threads
    # conex_max_ssh = 10
    # >=10 apresentara falha na conexão com o SSH devido limite
    # 8 threads melhor resposta
    hosts_lists = create_batch()
    collected_results = []
    all_results = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=n_threads) as executor:
        ssh_client = ssh_connect()
        for host_list in hosts_lists:
            results = {}
            futures = []
            for host in host_list:
                snmp_host = host['ip']
                snmp_oid = '1.3.6.1.2.1.17.4.3.1.1'  # mac
                snmp_oid2 = '1.3.6.1.2.1.17.4.3.1.2'  # porta
                ##Executor
                result = executor.submit(snmpwalk, snmp_host, snmp_oid, snmp_oid2, ssh_client, thread_lock)
                futures.append(result)
                for future in concurrent.futures.as_completed(futures):
                    results[host['host']] = {
                        'Nome host': host['host'],
                        'ip': host['ip'],
                        'snmp_result': future.result()  # Mantenha o future aqui
                    }

                # print(results)
                # print(results['snmp_mac_port_result'])



            # Aguarde todos os resultados terminarem e colete-os
            for host, result_data in results.items():
                result_mac, result_port = result_data['snmp_result']
                mac_port_list = {'mac': result_mac.split('\n'),
                                 'port': result_port.split('\n')}
                collected_results.append((host, result_data['ip'], mac_port_list))
        ssh_client.close()
    all_results.append(collected_results)
        # time.sleep(20)
    # ssh_client.close()
    print(f"desconectado do ssh")
    return collected_results,all_results

def create_batch(): #Teste para criação de lotes de solicitação
    list = extract_xlxs()
    sub_list = []
    lote = 100

    for i in range(0, len(list), lote):
        sublista = list[i:i + lote]
        sub_list.append(sublista)
    print(sub_list)
    for lista in sub_list:
        print (lista)
    return sub_list



def main(search_mac=''):
    start_time = time.time()
    snmpwalk_result, all_result = thread_execute()
    end_time = time.time()
    execution_time = end_time - start_time
    mac_identifier = []

    print("resultados:\n")
    for host, ip, mac_port_data in snmpwalk_result:

        print(f"Nome do Host: {host}")
        print(f"IP: {ip}")

        mac_list = mac_port_data['mac']
        port_list = mac_port_data['port']

        valid_entries = [(mac.strip(), port.strip()) for mac, port in zip(mac_list, port_list) if
                         mac.strip() and port.strip()]

        for mac, port in valid_entries:
            print(f"MAC: {mac} - Porta: {port}")
            # if mac == search_mac:
            #     var = {
            #         'nome do host': host,
            #         ''
            #     }
            #     mac_identifier.append

        print("==" * 30)
    save_excel(snmpwalk_result)
    print(all_result)
    print(f"Tempo de execução snmpwalk: {execution_time:.2f} segundos")



if __name__ == '__main__':
    # start_time = time.time()
    main()
# #schedule config
# import schedule
# schedule.every(5).minutes.do(JOB)
#
# while True:
#     schedule.run_pending()
#     time.sleep(1)